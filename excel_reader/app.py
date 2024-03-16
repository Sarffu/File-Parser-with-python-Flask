from flask import Flask, jsonify, request, render_template
from flask_marshmallow import Marshmallow
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook  # to read /write at excel file..

app = Flask(__name__)

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///book.db"
db = SQLAlchemy(app)
ma = Marshmallow(app)  # ye mm obj hai for serialize & deserialise flask instance....


class Parser(db.Model):  # ye sab variables ke sath connnect krr rhaa h database me..
    b_no = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(30), nullable=False)
    author = db.Column(db.String(50), nullable=False)


class Parserschema(ma.Schema):
    class Meta:
        fields = ["b_no", "name", "author"]


parser_schema = Parserschema()
parser_schemas = Parserschema(many=True)


# to fetch books ...
@app.route("/", methods=["GET"])
def home():
    return render_template("index.html")


#  for add books in db &  file xception handling...
@app.route("/add", methods=["POST"])
def add_data():

    ec_data = request.files["MyBooks"]
    if ec_data.filename == " ":
        return "No selected file"

    try:
        if ec_data.filename.endswith(
            ".xlsx"
        ):  # checks extension xls file me h ya nhiii.
            MyBooks = load_workbook(ec_data)
            NewBook = MyBooks.active

            for i in NewBook.iter_rows(
                min_row=2, values_only=True
            ):  # row = 2 means first row title ka hona chahiye & second row se values  hona chahiye..

                if None not in i:  # checks any empty cell or not in sheet..
                    book = Parser(name=i[0], author=i[1])
                    db.session.add(book)
                    db.session.commit()
            return "Message : Data retrieved..."
        else:
            return "Invalid file format. Please upload a file with Excel file format.."
    except Exception as e:
        return f"Error: {str(e)}"  # if any error aaye to Error msg show krega...


@app.route("/get", methods=["GET"])
def fetch_data():
    all_posts = Parser.query.all()  # all data fetch  ho rha h using parser model..
    result = parser_schemas.dump(
        all_posts
    )  # yahhaa mm kaa serialze ho rha h over network..
    return jsonify(result)


# get single book by id..
@app.route("/get/<int:b_no>", methods=["GET"])
def get_data(b_no):
    post = Parser.query.filter_by(
        b_no=b_no
    ).first()  # first jo h first result show kregaa...
    if post:
        result = parser_schema.dump(
            post
        )  # yaha post obj serialises into json serialbl format
        return jsonify(result)
    else:
        return jsonify({"message": "Book not found"}), 404


#  update existing book..
@app.route("/update/<int:b_no>", methods=["PUT"])
def update_book(b_no):
    post = Parser.query.get(b_no)
    if not post:
        return jsonify({"message": "Book not found"})

    data = request.get_json()
    post.name = data.get("name", post.name)
    post.author = data.get("author", post.author)

    db.session.commit()
    return parser_schema.jsonify(post)


#  deleting book from db....
@app.route("/delete/<int:b_no>", methods=["DELETE"])
def delete_book(b_no):
    post = Parser.query.get(b_no)
    if not post:
        return jsonify({"message": "Book not found"})

    db.session.delete(post)
    db.session.commit()
    return parser_schema.jsonify(post)


if __name__ == "__main__":
    app.run(debug=True, port=8000)
